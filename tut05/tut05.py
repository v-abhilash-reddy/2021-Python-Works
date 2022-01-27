import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
#os.system('cls')

def generate_marksheet():
    if not os.path.exists("output"):
        os.mkdir("output")
#gradepoints and heading 
    grade_points = {'AA':10,'AB':9,'BB':8,'BC':7,'CC':6,'CD':5,'DD':4,'F':0,'I':0,'AA*':10,'AB*':9,'BB*':8,'BC*':7,'CC*':6,'CD*':5,'DD*':4,'F*':0,'I*':0}
    heading = ["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"]
#opening files
    names_roll = open('names-roll.csv', 'r')
    grades = open('grades.csv', 'r')
    subs_master = open( 'subjects_master.csv','r')
#reading files into variables
    names_roll_reader = csv.reader(names_roll)
    grades_reader = csv.reader(grades)
    subs_reader = csv.reader(subs_master)
#listing
    listing_roll = [list(row) for row in names_roll_reader]
    listing_grades = [list(row) for row in grades_reader]
    listing_subs =[list(row) for row in subs_reader]
#taking subjects as a dictionary
    dict_for_subs = {}
    for sub in listing_subs[1:]:
        if sub[0] not in dict_for_subs:
            dict_for_subs[sub[0]] = sub
#creation of excel marksheet and 9 subsheets for each student
    for roll_no in listing_roll[1:]:
        spi,credits = [],[]
        wb = Workbook()
        for i in range(1,9):
            sumofcreds=0
            wgtedsum=0
            if f'Sem{i}' not in wb.sheetnames:
                wb.create_sheet(f'Sem{i}',i)
            ws = wb[f"Sem{i}"]
            ws.append(heading)
            j = 1
            for row in listing_grades[1:]:
                if row[0]==roll_no[0] and int(row[1])==i :
                    info = []
                    info.append(j); j += 1
                    info.append(row[2])
                    var = dict_for_subs[row[2]]        
                    info.append(var[1])
                    info.append(var[2])
                    info.append(var[3])
                    sumofcreds += int(row[3].strip())
                    info.append(row[5])
                    info.append(row[4])
                    wgtedsum += grade_points[f'{row[4].strip()}']*int(row[3])
                    ws.append(info)   
            if sumofcreds!=0 :
                spi.append(round((wgtedsum/sumofcreds),2))
                credits.append(sumofcreds)
        ws = wb.active
        ws.title = "Overall"
        ws.append(["RollNo",roll_no[0]])
        ws.append(["Name of Student",roll_no[1]])
        ws.append(["Discipline",roll_no[0][4:6]])
        ws.append(["Semester No"]+[x for x in range(1,i)])
        ws.append(["Semester wise Credit Taken"]+credits)
        ws.append(["SPI"]+spi)
        ws.append(["Total Credits taken"]+[sum(credits[:x+1]) for x in range(len(credits))])

        creditsxspi =0; credits_sum = 0
        all_cpi = []
        for j in range(len(credits)):
            credits_sum += credits[j]
            creditsxspi += spi[j]*credits[j]
            cpi = round(creditsxspi/credits_sum,2)
            all_cpi.append(cpi)
        ws.append(["CPI"]+all_cpi)
        wb.save(f'output\\{roll_no[0]}.xlsx')
    return

#function that generates marksheet of each student
generate_marksheet()