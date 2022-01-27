import os,math,csv,openpyxl,pandas as pd
from numpy import NaN
from openpyxl.styles.colors import BLUE
from openpyxl.styles import Alignment
from openpyxl.styles import Font,Color
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side

sample_output_path = "./sample_output/marksheet"

def get_roll_email():
    slist = pd.read_csv('./sample_input/master_roll.csv')
    response = pd.read_csv('./sample_input/responses.csv')
    lst = []
    for index,row in response.iterrows():
        lst.append([row["Roll Number"],row["Email address"],row["IITP webmail"]])
    return lst

def get_answer():
    slist = pd.read_csv('./sample_input/master_roll.csv')
    response = pd.read_csv('./sample_input/responses.csv')
    for index,row in response.iterrows():
        if row['Roll Number'] == 'ANSWER':
            return [key for key in row][7:]
    return []

def calculate(mrks,wmrks,crt_opts_list,curr_opts):
    crt_opts,not_attempted,wrg_opts,lt,r = 0,0,0,[],0
    for j in range(len(crt_opts_list)):
        if curr_opts[j] == crt_opts_list[j]: crt_opts+=1
        else : wrg_opts+=1
        r+=1
        lt.append(curr_opts[j])
    not_attempted = pd.Series(lt).isna().sum()
    wrg_opts-=not_attempted
    result = crt_opts*mrks + wrg_opts*wmrks
    return result,[crt_opts,wrg_opts,not_attempted],crt_opts*mrks

def generate_concise(mrks,wmrks):
    slist = pd.read_csv('./sample_input/master_roll.csv')
    response = pd.read_csv('./sample_input/responses.csv')
    roll_dict = {}
    for i in range(len(response)):
        roll_dict[response.at[i,"Roll Number"].upper()] = 1
    for i in range(len(slist)):
        if slist.at[i,"roll"] in roll_dict: continue
        else: 
            length = len(response)
            cols = response.columns
            lst = []
            for item in cols:
                print(item)
                if item == "Roll Number":
                    lst.append(slist.at[i,"roll"])
                elif item == "Name":
                    lst.append(slist.at[i,"name"])
                else :
                    lst.append(NaN)
            response.loc[length] = lst 
    concise_marksheet = response
    crt_opts_list = get_answer()
    if crt_opts_list == []:
        return "Error!!!Answer is not exist in responses"
    
    last_list,score_af_neg,score_bf_neg= [],[],[]
    for index,row in response.iterrows():
        curr_opts = [key for key in row][7:]
        result,lst,oc = calculate(mrks,wmrks,crt_opts_list,curr_opts)
        last_list.append(lst)
        score_af_neg.append(str(round(result,2))+"/"+str(mrks*len(crt_opts_list)))
        score_bf_neg.append(str(round(oc,2))+"/"+str(mrks*len(crt_opts_list)))
    concise_marksheet.insert(loc =6,column ="Score_After_Negative",value =score_af_neg)
    concise_marksheet.insert(loc=6,column ="Score_before_Negative",value =score_bf_neg)
    concise_marksheet["Options"] = last_list
    os.makedirs("sample_output",exist_ok = True)
    os.makedirs(sample_output_path,exist_ok = True)
    concise_marksheet.to_csv(sample_output_path+"/concise_marksheet.csv", index=False)
    return  
   




   
def generateMarksheet(mrks,wmrks):
    slist = pd.read_csv('./sample_input/master_roll.csv')   
    response = pd.read_csv('./sample_input/responses.csv')
    crt = get_answer()
    if crt == []:
        return "Error!!!Answer sheet does not exist in the responses"
    #edge case or absentees
    roll_dict = {}
    for i in range(len(response)):
        roll_dict[response.at[i,"Roll Number"].upper()] = 1
    
    # n = 1
    for i in range(len(slist)):
        if slist.at[i,"roll"] in roll_dict: continue
        else: 
            length = len(response)
            cols = response.columns
            lst = []
            for item in cols:
                print(item)
                if item == "Roll Number":
                    lst.append(slist.at[i,"roll"])
                elif item == "Name":
                    lst.append(slist.at[i,"name"])
                else :
                    lst.append(NaN)
            response.loc[length] = lst
            # n+=1
        
    os.makedirs("sample_output",exist_ok = True)
    os.makedirs(sample_output_path,exist_ok = True)
    for i in range(len(response)):       
        #creation and initialising a workbook
        wb =  openpyxl.Workbook()
        sheet = wb.active
        #increase cells width
        sheet.column_dimensions['A'].width=17
        sheet.column_dimensions['B'].width=17
        sheet.column_dimensions['C'].width=17
        sheet.column_dimensions['D'].width=17
        sheet.column_dimensions['E'].width=17
        for z in range(1,41): sheet.row_dimensions[z].height = 20

        #alignment to centre
        alignment = Alignment(horizontal='center',vertical='bottom',text_rotation=0,wrap_text=False,shrink_to_fit=True,indent=0)
        
        sheet.title = response.at[i,'Roll Number'].upper()
        sheet.merge_cells('A1:E4')
        # add image here
        img = Image('./static/iitp_logo.png')
        img.width=610
        img.height=80
        sheet.add_image(img,'A1')

        sheet.merge_cells('A5:E5')
        sheet.cell(row=5,column=1).value = "Mark Sheet"
        sheet.cell(row=5,column=1).alignment = alignment
        sheet.cell(row=5,column=1).font = Font(size=18,bold=True,name='Century',underline='single')

        #borders
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),  top=Side(style='thin'), bottom=Side(style='thin'))
        rn,cn = 9,1
        while rn!=13:
            if(cn==6): cn=1;rn+=1
            sheet.cell(row=rn,column=cn).border = thin_border
            cn+=1 

        sheet.merge_cells('B6:C6')
        flist = [[6,1,"Name:"],[6,2,slist.at[i,'name']],[6,4,"Exam:"],[6,5,"quiz"],[7,1,"Roll Number:"],[7,2,response.at[i,'Roll Number'].upper()],[9,2,"Right"],[9,3,"Wrong"],[9,4,"Not Attempt"],[9,5,"Max"],[10,1,"No."],[11,1,"Marking"],[12,1,"Total"]]
        for s in flist:
            sheet.cell(row=s[0],column=s[1]).value = s[2]
            sheet.cell(row=s[0],column=s[1]).alignment = alignment
            if(s[2]!="Name:" and s[2]!="Exam:" and s[2]!="Roll Number:"): sheet.cell(row=s[0],column=s[1]).font = Font(bold=True,size=12)

        sheet.cell(row=10,column=2).font = Font(color="00008000")
        sheet.cell(row=11,column=2).font = Font(color="00008000")
        sheet.cell(row=12,column=2).font = Font(color="00008000")
        sheet.cell(row=10,column=3).font = Font(color="00FF0000")
        sheet.cell(row=11,column=3).font = Font(color="00FF0000")
        sheet.cell(row=12,column=3).font = Font(color="00FF0000")
        sheet.cell(row=12,column=5).font = Font(color=BLUE)
        
        rgt,wrg,na=0,0,0
        sheet.cell(row=15,column=1).value = "Student Ans"
        sheet.cell(row=15,column=1).alignment = alignment
        sheet.cell(row=15,column=1).border = thin_border
        sheet.cell(row=15,column=1).font = Font(bold=True,size=12)
        sheet.cell(row=15,column=2).value = "Correct Ans"
        sheet.cell(row=15,column=2).alignment = alignment
        sheet.cell(row=15,column=2).border = thin_border
        sheet.cell(row=15,column=2).font = Font(bold=True,size=12)
        sheet.cell(row=15,column=4).value = "Student Ans"
        sheet.cell(row=15,column=4).alignment = alignment
        sheet.cell(row=15,column=4).border = thin_border
        sheet.cell(row=15,column=4).font = Font(bold=True,size=12)
        sheet.cell(row=15,column=5).value = "Correct Ans"
        sheet.cell(row=15,column=5).alignment = alignment
        sheet.cell(row=15,column=5).border = thin_border
        sheet.cell(row=15,column=5).font = Font(bold=True,size=12)

        ## create empty files when they are not submitted
        lt = []
        r,c = 16,1
        for j in range(7,7+len(crt)) : #Checking the Options
            if(j==32): r,c=16,4
            sheet.cell(row=r,column=c).value = response.iat[i,j]
            sheet.cell(row=r,column=c).alignment = alignment
            sheet.cell(row=r,column=c).border = thin_border
            if(response.iat[i,j]==crt[j-7]): sheet.cell(row=r,column=c).font = Font(color="00008000")
            else : sheet.cell(row=r,column=c).font = Font(color="00FF0000")
            sheet.cell(row=r,column=c+1).value = crt[j-7]
            sheet.cell(row=r,column=c+1).alignment = alignment
            sheet.cell(row=r,column=c+1).border = thin_border
            sheet.cell(row=r,column=c+1).font = Font(color=BLUE)
            if(response.iat[i,j]==crt[j-7]):#counting right,wrong and not attempted answers 
                rgt+=1
            else: wrg+=1
            r+=1
            lt.append(response.iat[i,j])

        na = pd.Series(lt).isna().sum()
        wrg-=na
        
        seclist = [[10,2,rgt],[11,2,mrks],[12,2,rgt*mrks],[10,3,wrg],[11,3,wmrks],[12,3,wrg*wmrks],[10,4,na],[11,4,0],[10,5,rgt+wrg+na],[12,5,str(round(rgt*mrks+wrg*wmrks,2))+'/'+str(len(crt)*mrks)]]
        for s in seclist:
            sheet.cell(row=s[0],column=s[1]).value = s[2]
            sheet.cell(row=s[0],column=s[1]).alignment = alignment
        
        wb.save("./sample_output/marksheet/"+sheet.title+'.xlsx')
    return

# generateMarksheet(5,-1)