import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
def output_individual_roll():
    data = open("regtable_old.csv", "r")
    data_csv=csv.reader(data)
    data_list =[]
    for row in data_csv:
        data_list.append(list(row))
    if not os.path.exists("output_individual_roll"):
        os.mkdir("output_individual_roll")
    for row in data_list[1:]:
        roll_no = row[0]
        register_sem = row[1]
        sub_no = row[3]
        sub_type = row[-1]
        rno='{}.xlsx'.format(roll_no)
        if(os.path.isfile('./output_individual_roll/'+rno)): 
            workbook=load_workbook(r'output_individual_roll\\{}.xlsx'.format(roll_no))
            sheet=workbook.active
            sheet.append([ roll_no, register_sem,sub_no,sub_type])
            workbook.save(r'output_individual_roll\\{}.xlsx'.format(roll_no))   
        else: 
            workbook=Workbook()
            sheet=workbook.active
            sheet.append(["rollno","register_sem","subno","sub_type"])
            sheet.append([ roll_no, register_sem,sub_no,sub_type])
            workbook.save(f'output_individual_roll\\{roll_no}.xlsx')
    return



def output_by_subject():
    data= open("regtable_old.csv", "r")
    data_csv=csv.reader(data)
    data_list =[]
    for row in data_csv:
        data_list.append(list(row))
    if not os.path.exists("output_by_subject"):
        os.mkdir("output_by_subject")
    for row in data_list[1:]:
        roll_no = row[0]
        register_sem = row[1]
        sub_no = row[3]
        sub_type = row[-1]
        sno='{}.xlsx'.format(sub_no)
        path='./output_by_subject/'+sno
        if(os.path.isfile(path)): 
            workbook=load_workbook(r'output_by_subject\\{}.xlsx'.format(sub_no))
            sheet=workbook.active
            sheet.append([ roll_no, register_sem,sub_no,sub_type])
            workbook.save(r'output_by_subject\\{}.xlsx'.format(sub_no))   
        else: 
            workbook=Workbook()
            sheet=workbook.active
            sheet.append(["rollno","register_sem","subno","sub_type"])
            sheet.append([ roll_no, register_sem,sub_no,sub_type])
            workbook.save(f'output_by_subject\\{sub_no}.xlsx')  
    return

output_individual_roll()
output_by_subject()