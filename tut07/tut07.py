import pandas as pd
import openpyxl

def feedback_not_submitted():

	ltp_mapping_feedback_type = {1: 'lecture', 2: 'tutorial', 3:'practical'}
	output_file_name = "course_feedback_remaining.xlsx" 
	
	crbas = pd.read_csv('./course_registered_by_all_students.csv')#reading all the csv files into variables
	cm = pd.read_csv('./course_master_dont_open_in_excel.csv')
	cfsbs = pd.read_csv('./course_feedback_submitted_by_students.csv')
	si = pd.read_csv('./studentinfo.csv')
	
	wb =  openpyxl.Workbook() #creation and initialising a workbook
	sheet = wb.active
	sheet.title = 'feedback'
	s = ['rollno','register_sem','schedule_sem','subno','name','email','aemail','contact']
	for col in range(1,9): sheet.cell(row=1,column=col).value = s[col-1]
	ro = 2

	cm_dict = {}
	for j in range(len(cm)):
		cm_dict[cm.at[j,'subno']] = cm.at[j,'ltp']

	cfsbs_dict = {}
	for j in range(len(cfsbs)):
		cfsbs_dict[cfsbs.at[j,'stud_roll']+cfsbs.at[j,'course_code']+str(cfsbs.at[j,'feedback_type'])] = 1

	si_dict = {}
	for j in range(len(si)):
		si_dict[si.at[j,'Roll No']] = j	

	for i in range(len(crbas)):#looping for the search of students who did not fill the feedback
		roll = crbas.at[i,'rollno']
		course = crbas.at[i,'subno']

		ltp = cm_dict[course]
		x = ltp.split('-')
		l = '1' if int(x[0]) else ''
		t = '2' if int(x[1]) else ''
		p = '3' if int(x[2]) else ''
		
		a = 1
		if l:
			if roll+course+l in cfsbs_dict.keys(): pass
			else : a = 0
		if t:
			if roll+course+t in cfsbs_dict.keys(): pass
			else : a = 0
		if p:
			if roll+course+p in cfsbs_dict.keys(): pass
			else : a = 0

		if not a:
			if roll in si_dict.keys():
				y = si_dict[roll]
				z = [roll,crbas.at[i,'register_sem'],crbas.at[i,'schedule_sem'],course,si.at[y,'Name'],si.at[y,'email'],si.at[y,'aemail'],si.at[y,'contact']]
				for col in range(1,9): sheet.cell(row=ro,column=col).value = z[col-1]
			else:
				z = [roll,crbas.at[i,'register_sem'],crbas.at[i,'schedule_sem'],course,"NA_IN_STUDENTINFO","NA_IN_STUDENTINFO","NA_IN_STUDENTINFO","NA_IN_STUDENTINFO"]
			for col in range(1,9): sheet.cell(row=ro,column=col).value = z[col-1]
			ro += 1

	wb.save(output_file_name)


feedback_not_submitted()